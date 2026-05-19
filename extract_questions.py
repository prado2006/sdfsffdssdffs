from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX = Path(r"C:/Users/Asort/Downloads/Вопросы_по_странам (1).xlsx")
OUTPUT_JSON = ROOT / "data" / "questions.json"

ROUTE = [
    "Красноярск",
    "Монголия",
    "Китай",
    "Мьянма",
    "Таиланд",
    "Индонезия",
    "Филиппины",
    "Перу",
    "Бразилия",
    "Камерун",
    "Уганда",
    "Сомали",
    "Индия",
    "Узбекистан",
    "Красноярск",
]


def cell_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def clean_answer(value: str) -> str:
    return re.sub(r"\s*\([^)]*\)", "", value).strip()


def main() -> None:
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    sheet = workbook.active

    current_block = ""
    questions: list[dict[str, Any]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        number = cell_text(row[0])
        if number.lower().startswith("блок"):
            current_block = number
            continue

        country = cell_text(row[1])
        question = cell_text(row[2])
        answers = [clean_answer(cell_text(row[index])) for index in range(3, 7)]
        correct = clean_answer(cell_text(row[7]))

        if not country or not question or not correct:
            continue

        if country.startswith("Россия"):
            country_key = "Россия"
        else:
            country_key = country

        questions.append(
            {
                "id": len(questions) + 1,
                "block": current_block,
                "source_number": number,
                "country": country_key,
                "question": question,
                "answers": answers,
                "correct": correct,
            }
        )

    payload = {
        "title": "Вокруг света за 80 дней",
        "route": ROUTE,
        "questions": questions,
    }

    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"Saved {len(questions)} questions to {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
